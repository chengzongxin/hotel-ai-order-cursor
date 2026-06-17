from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config.settings import settings


@dataclass(frozen=True)
class ServiceProductRecord:
    service_product_code: str
    service_product_name: str
    product_type: str
    category: str
    raw_service_type: str
    service_order_type: str
    unit: str
    price: str
    price_status: str
    shelf_status: str
    repair_category: str
    related_category: str
    related_area: str
    fault_phenomenon: str
    display_order: str
    remark: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class SpuExcelLoader:
    def __init__(self, excel_path: str | Path | None = None) -> None:
        self.excel_path = Path(excel_path or settings.spu_excel_path)

    def load(self) -> list[ServiceProductRecord]:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"SPU Excel not found: {self.excel_path}")

        workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [self._to_text(cell) for cell in next(rows)]

        records: list[ServiceProductRecord] = []
        for row in rows:
            item = dict(zip(headers, row, strict=False))
            record = ServiceProductRecord(
                service_product_code=self._to_text(item.get("服务商品编码")),
                service_product_name=self._to_text(item.get("服务商品名称")),
                product_type=self._to_text(item.get("商品类型")),
                category=self._to_text(item.get("所属分类")),
                raw_service_type=self._to_text(item.get("所属服务类型")),
                service_order_type=self._normalize_service_order_type(
                    product_type=self._to_text(item.get("商品类型")),
                    raw_service_type=self._to_text(item.get("所属服务类型")),
                    name=self._to_text(item.get("服务商品名称")),
                    remark=self._to_text(item.get("备注")),
                ),
                unit=self._to_text(item.get("计量单位")),
                price=self._to_text(item.get("商品实际价格")),
                price_status=self._to_text(item.get("商品价格状态")),
                shelf_status=self._to_text(item.get("上下架状态")),
                repair_category=self._to_text(item.get("维修分类")),
                related_category=self._to_text(item.get("关联品类")),
                related_area=self._to_text(item.get("关联区域")),
                fault_phenomenon=self._to_text(item.get("关联故障现象")),
                display_order=self._to_text(item.get("显示顺序")),
                remark=self._to_text(item.get("备注")),
            )
            if record.service_product_code and record.service_product_name and record.shelf_status != "下架":
                records.append(record)

        return records

    def _normalize_service_order_type(
        self,
        product_type: str,
        raw_service_type: str,
        name: str,
        remark: str,
    ) -> str:
        text = f"{product_type} {raw_service_type} {name} {remark}"
        if "托管" in text:
            return "托管维修"
        if "测量" in text or "量尺" in text or "量房" in text:
            return "单次测量"
        if "安装" in text or "拆装" in text:
            return "单次安装"
        if "维修" in text or "修" in text:
            return "单次维修服务"
        return raw_service_type or product_type

    def _to_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()
