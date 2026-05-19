import hashlib
import json
from dataclasses import asdict, dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

from config.settings import settings


@dataclass(frozen=True)
class ProductRecord:
    code: str
    name: str
    product_type: str
    category: str
    service_type: str
    unit: str
    price: str
    price_status: str
    shelf_status: str
    repair_category: str
    related_category: str
    related_area: str
    fault_phenomenon: str
    remark: str


@dataclass(frozen=True)
class RecallResult:
    score: float
    record: ProductRecord

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self.record)
        payload["score"] = round(float(self.score), 4)
        return payload


class ProductEmbeddingRetriever:
    """维修商品 embedding 召回器。

    工作流程：
    1. 从 Excel 读取商品和故障文本。
    2. 用 sentence-transformers 生成向量。
    3. 用 numpy 计算 cosine similarity。
    4. 根据 threshold 过滤低相关结果。
    """

    def __init__(
        self,
        excel_path: str | Path | None = None,
        model_name: str | None = None,
        cache_dir: str | Path | None = None,
    ) -> None:
        self.excel_path = Path(excel_path or settings.spu_excel_path)
        self.model_name = model_name or settings.embedding_model_name
        self.cache_dir = Path(cache_dir or settings.embedding_cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)

        self._model: Any | None = None
        self._records: list[ProductRecord] | None = None
        self._product_embeddings: np.ndarray | None = None
        self._fault_embeddings: np.ndarray | None = None

    def search_products(
        self,
        query: str,
        top_k: int = 5,
        threshold: float | None = None,
    ) -> list[dict[str, Any]]:
        results = self._search(
            query=query,
            embeddings=self.product_embeddings,
            top_k=top_k,
            threshold=threshold if threshold is not None else settings.product_recall_threshold,
        )
        return [result.to_dict() for result in results]

    def search_faults(
        self,
        query: str,
        top_k: int = 5,
        threshold: float | None = None,
    ) -> list[dict[str, Any]]:
        results = self._search(
            query=query,
            embeddings=self.fault_embeddings,
            top_k=top_k,
            threshold=threshold if threshold is not None else settings.fault_recall_threshold,
        )
        return [result.to_dict() for result in results]

    @property
    def records(self) -> list[ProductRecord]:
        if self._records is None:
            self._records = self._load_records()
        return self._records

    @property
    def product_embeddings(self) -> np.ndarray:
        if self._product_embeddings is None:
            self._product_embeddings = self._load_or_build_embeddings("product")
        return self._product_embeddings

    @property
    def fault_embeddings(self) -> np.ndarray:
        if self._fault_embeddings is None:
            self._fault_embeddings = self._load_or_build_embeddings("fault")
        return self._fault_embeddings

    @property
    def model(self) -> Any:
        if self._model is None:
            from sentence_transformers import SentenceTransformer

            self._model = SentenceTransformer(self.model_name)
        return self._model

    def _search(
        self,
        query: str,
        embeddings: np.ndarray,
        top_k: int,
        threshold: float,
    ) -> list[RecallResult]:
        clean_query = query.strip()
        if not clean_query:
            return []

        query_embedding = self._normalize(
            self.model.encode([clean_query], convert_to_numpy=True)
        )[0]
        scores = embeddings @ query_embedding
        ranked_indexes = np.argsort(scores)[::-1]

        results: list[RecallResult] = []
        for index in ranked_indexes:
            score = float(scores[index])
            if score < threshold:
                continue
            results.append(RecallResult(score=score, record=self.records[int(index)]))
            if len(results) >= top_k:
                break

        return results

    def _load_or_build_embeddings(self, kind: str) -> np.ndarray:
        cache_path = self._cache_path(kind)
        metadata_path = cache_path.with_suffix(".json")
        current_metadata = self._cache_metadata(kind)

        if cache_path.exists() and metadata_path.exists():
            cached_metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            if cached_metadata == current_metadata:
                return np.load(cache_path)

        texts = [
            self._build_product_text(record) if kind == "product" else self._build_fault_text(record)
            for record in self.records
        ]
        embeddings = self.model.encode(
            texts,
            batch_size=32,
            convert_to_numpy=True,
            show_progress_bar=False,
        )
        normalized_embeddings = self._normalize(embeddings)

        np.save(cache_path, normalized_embeddings)
        metadata_path.write_text(
            json.dumps(current_metadata, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return normalized_embeddings

    def _load_records(self) -> list[ProductRecord]:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"SPU Excel not found: {self.excel_path}")

        workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [self._to_text(cell) for cell in next(rows)]

        records: list[ProductRecord] = []
        for row in rows:
            item = dict(zip(headers, row, strict=False))
            record = ProductRecord(
                code=self._to_text(item.get("服务商品编码")),
                name=self._to_text(item.get("服务商品名称")),
                product_type=self._to_text(item.get("商品类型")),
                category=self._to_text(item.get("所属分类")),
                service_type=self._to_text(item.get("所属服务类型")),
                unit=self._to_text(item.get("计量单位")),
                price=self._to_text(item.get("商品实际价格")),
                price_status=self._to_text(item.get("商品价格状态")),
                shelf_status=self._to_text(item.get("上下架状态")),
                repair_category=self._to_text(item.get("维修分类")),
                related_category=self._to_text(item.get("关联品类")),
                related_area=self._to_text(item.get("关联区域")),
                fault_phenomenon=self._to_text(item.get("关联故障现象")),
                remark=self._to_text(item.get("备注")),
            )
            if record.name and record.shelf_status != "下架":
                records.append(record)

        return records

    def _build_product_text(self, record: ProductRecord) -> str:
        return self._join_text(
            [
                record.name,
                record.product_type,
                record.category,
                record.service_type,
                record.repair_category,
                record.related_category,
                record.related_area,
                record.remark,
            ]
        )

    def _build_fault_text(self, record: ProductRecord) -> str:
        return self._join_text(
            [
                record.name,
                record.category,
                record.repair_category,
                record.related_category,
                record.related_area,
                record.fault_phenomenon,
                record.remark,
            ]
        )

    def _cache_path(self, kind: str) -> Path:
        digest = hashlib.sha256(
            f"{self.excel_path.resolve()}:{self.model_name}:{kind}".encode("utf-8")
        ).hexdigest()[:16]
        return self.cache_dir / f"spu_{kind}_{digest}.npy"

    def _cache_metadata(self, kind: str) -> dict[str, Any]:
        stat = self.excel_path.stat()
        return {
            "kind": kind,
            "excel_path": str(self.excel_path.resolve()),
            "excel_mtime": stat.st_mtime,
            "excel_size": stat.st_size,
            "model_name": self.model_name,
            "record_count": len(self.records),
        }

    def _normalize(self, embeddings: np.ndarray) -> np.ndarray:
        norms = np.linalg.norm(embeddings, axis=1, keepdims=True)
        norms[norms == 0] = 1
        return embeddings / norms

    def _join_text(self, values: list[str]) -> str:
        return " ".join(value for value in values if value)

    def _to_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()


@lru_cache
def get_product_retriever() -> ProductEmbeddingRetriever:
    return ProductEmbeddingRetriever()
